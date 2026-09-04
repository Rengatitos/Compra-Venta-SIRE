"""Carga del maestro de cuentas desde el Excel que exporta Contasis.

Aquí vive lo que toca openpyxl; la interpretación de las filas es de
`app.domain.plan_cuentas`, que no depende de ninguna librería de Excel.
"""

from __future__ import annotations

import io
import logging
import zipfile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.domain.plan_cuentas import NOMBRE_HOJA, Cuenta, localizar_columnas, parsear

logger = logging.getLogger(__name__)


class ExcelInvalido(ValueError):
    """El archivo no es un maestro de cuentas legible.

    Se distingue de un error interno porque tiene arreglo del lado del
    usuario: subir el archivo correcto.
    """


def _hoja(wb):
    """La hoja del maestro, buscada por nombre y con respaldo por contenido.

    Contasis exporta la hoja como «PLAN DE CUENTAS», pero un archivo que haya
    pasado por las manos de alguien puede traerla renombrada. En ese caso se
    busca la primera que tenga la fila de cabeceras esperada, que es una señal
    más fiable que el nombre.
    """
    for nombre in wb.sheetnames:
        if nombre.strip().upper() == NOMBRE_HOJA:
            return wb[nombre]

    for nombre in wb.sheetnames:
        hoja = wb[nombre]
        for fila in hoja.iter_rows(max_row=20, values_only=True):
            if localizar_columnas(fila):
                logger.info("Maestro de cuentas encontrado en la hoja «%s»", nombre)
                return hoja

    raise ExcelInvalido(
        f"El archivo no tiene una hoja «{NOMBRE_HOJA}» ni ninguna con las columnas "
        "CUENTA y DESCRIPCION"
    )


def desde_excel(contenido: bytes) -> list[Cuenta]:
    """Cuentas del archivo subido.

    `read_only` importa: son casi tres mil filas y el modo normal las carga
    todas como objetos de celda.
    """
    if not contenido:
        raise ExcelInvalido("El archivo llegó vacío")

    try:
        wb = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    except (
        InvalidFileException,
        zipfile.BadZipFile,
        KeyError,
        OSError,
        ValueError,
    ) as fallo:
        raise ExcelInvalido(f"No se pudo abrir el archivo como Excel: {fallo}") from fallo

    try:
        cuentas = parsear(_hoja(wb).iter_rows(values_only=True))
    except ValueError as fallo:
        raise ExcelInvalido(str(fallo)) from fallo
    finally:
        wb.close()

    if not cuentas:
        raise ExcelInvalido("La hoja no tiene ninguna cuenta con código utilizable")

    return cuentas
