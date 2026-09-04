"""Exportación al formato oficial «REGISTRO DE COMPRAS Y VENTAS» de Contasis.

La plantilla vive en `app/resources/plantilla_registro.xlsx` (generada por
[scripts/preparar_plantilla.py](../../scripts/preparar_plantilla.py) a partir
de `source/PLANTILLA REGISTRO DE COMPRAS Y VENTAS.xlsx`, que está fuera del
control de versiones y no entra a la imagen de Docker). No se reconstruye el
layout a mano: se abre el archivo real y se escriben las filas debajo de sus
encabezados, así el formato queda idéntico al que espera el contador.

El encabezado de tres niveles ocupa las filas 1-3 (sin las notas de uso, el
título ni la fila de especificación que trae la plantilla original de
Contasis: ningún registro real de contador las conserva) y los datos empiezan
en la fila 4, que además sirve de prototipo de estilo.

El registro se lleva en soles, como lo hacen los contadores: un comprobante en
dólares se convierte multiplicando por su tipo de cambio, y el importe
original en dólares va aparte, en la columna «EQUIVALENTE EN DOLARES
AMERICANOS». Ver `_Conversion` más abajo.

Del análisis IA viajan la cuenta contable y la glosa; si el clasificador RAG
(`app/services/ollama_rag.py`) llegó a una cuenta o glosa propias, ganan a las
del análisis general (ver `_rag`). El centro de costos no viaja nunca: la
plantilla pide el *código* del catálogo de Contasis (9 caracteres) y la IA
devuelve un nombre descriptivo, así que escribirlo recortado inventaría
códigos que colisionan entre sí.

Otras columnas se llenan con una regla en vez de con un dato de SUNAT, porque
así aparecen en el 100 % de los registros reales con los que se comparó esta
exportación: la condición de pago (CON/CRE, según haya vencimiento posterior a
la emisión), la cuenta contable total (4212 en compras, 1212 en ventas — o la
que haya resuelto el RAG para esta empresa) y el porcentaje de IGV (la tasa
declarada o, en su defecto, la general — en todas las filas, también las
exoneradas).
"""

from __future__ import annotations

import io
from copy import copy
from datetime import date
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Any, NamedTuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.comprobante import (
    Libro,
    normalizar_fecha,
    normalizar_tipo_cp,
)

RUTA_PLANTILLA = Path(__file__).resolve().parents[1] / "resources" / "plantilla_registro.xlsx"

# El encabezado de tres niveles ocupa las filas 1-3; los datos empiezan en la 4.
PRIMERA_FILA_DATOS = 4

# El nombre de la hoja de compras trae espacios de más: es literal dentro del
# archivo y openpyxl la busca tal cual.
HOJAS: dict[Libro, str] = {
    Libro.COMPRAS: "FORMATO_ COMPRAS ",
    Libro.VENTAS: "FORMATO_VENTAS",
}

FORMATO_FECHA = "DD/MM/YYYY"

# Contasis identifica la moneda con un solo carácter.
MONEDAS: dict[str, str] = {"PEN": "S", "USD": "D"}

# Tasa general. Se usa como respaldo cuando SUNAT no mandó la tasa del
# comprobante, tenga o no IGV: los cuatro registros reales con los que se
# comparó esta exportación llevan 18 en todas las filas, incluidas las
# operaciones exoneradas.
TASA_IGV = 18

# Anchos que declara la plantilla para las columnas de texto que llena la IA.
# Pasarse no es inofensivo: Contasis trunca por su cuenta al importar y el
# corte cae donde caiga.
MAX_CUENTA_CONTABLE = 10
MAX_GLOSA = 60

# Condición de pago (columna AE en compras, Y en ventas). Ninguna fuente de
# SUNAT la manda; se infiere de las fechas, como hace el contador: crédito
# cuando hay un vencimiento posterior a la emisión.
CONDICION_CONTADO = "CON"
CONDICION_CREDITO = "CRE"

# Cuenta contable del total del comprobante (columna AH en compras, AD en
# ventas), cuando el RAG no resolvió una propia para la empresa. Constante en
# los cuatro registros reales con los que se comparó esta exportación: PCGE
# 42.1.2 (proveedores) y 12.1.2 (clientes) — son también el valor por defecto
# que usa `ollama_rag.clasificar` cuando no encuentra un precedente.
CUENTA_TOTAL: dict[Libro, str] = {Libro.COMPRAS: "4212", Libro.VENTAS: "1212"}

_plantilla: bytes | None = None


def _bytes_plantilla() -> bytes:
    global _plantilla
    if _plantilla is None:
        _plantilla = RUTA_PLANTILLA.read_bytes()
    return _plantilla


def _texto(valor: Any) -> str | None:
    texto = "" if valor is None else str(valor).strip()
    return texto or None


def _entero_o_texto(valor: Any) -> int | str | None:
    """Serie y tipo van como texto; correlativos y documentos, como número.

    Es lo que hace el archivo de ejemplo de la plantilla: `E001` en la serie,
    pero `158` y `20612495522` como enteros.
    """
    texto = _texto(valor)
    if texto is None:
        return None
    return int(texto) if texto.isdigit() else texto


# El tipo de cambio no es un importe: la plantilla lo declara «(10,4)
# NUMERICO» y redondearlo a dos decimales convertía un 3.387 en 3.39.
DECIMALES_IMPORTE = 2
DECIMALES_TIPO_CAMBIO = 4


def _monto(valor: Any, factor: float = 1.0, decimales: int = DECIMALES_IMPORTE) -> float | None:
    """Un cero se escribe como celda vacía, no como `0`.

    Se redondea con `ROUND_HALF_UP` porque es como redondea un contador (y el
    resto del dominio, ver `normalizar_monto`): el `round()` de Python hace
    banker's rounding sobre binario y un `2.675` se guarda como `2.67`.
    `factor` convierte a soles un importe en moneda extranjera antes de
    redondear, para que la base y el IGV en soles sigan sumando el total.
    """
    try:
        numero = Decimal(str(float(valor))) * Decimal(str(factor))
    except (TypeError, ValueError):
        return None
    exponente = Decimal(1).scaleb(-decimales)
    return float(numero.quantize(exponente, rounding=ROUND_HALF_UP)) or None


def _suma(*valores: Any) -> float:
    total = 0.0
    for valor in valores:
        try:
            total += float(valor)
        except (TypeError, ValueError):
            continue
    return total


def _moneda(valor: Any) -> str:
    return MONEDAS.get(_texto(valor) or "PEN", "S")


def _recortar(valor: Any, tope: int) -> str | None:
    """Texto recortado a `tope`, cortando en el último espacio que quepa."""
    texto = _texto(valor)
    if texto is None or len(texto) <= tope:
        return texto
    corte = texto[:tope]
    espacio = corte.rfind(" ")
    # Sólo se corta por palabra si eso no deja un resto ridículo: con una
    # primera palabra larguísima es mejor el corte seco.
    if espacio > tope // 2:
        corte = corte[:espacio]
    return corte.rstrip(" ,;.-") or None


# El formato contable propio de la plantilla, sin símbolo de moneda: es el
# formato de todos los importes convertidos a soles y del pie de totales.
FORMATO_IMPORTE = r'_ * #,##0.00_ ;_ * \-#,##0.00_ ;_ * "-"??_ ;_ @_ '

# Sólo para la fila en moneda extranjera que no trae tipo de cambio: no se
# puede convertir a soles, así que se deja el importe nominal marcado con el
# símbolo de su moneda para que se note que no está en soles como el resto.
_FORMATO_US_DOLARES = r'_ * "US$"\ #,##0.00_ ;_ * "US$"\ \-#,##0.00_ ;_ * "-"??_ ;_ @_ '


def _formato_sin_convertir(moneda: str) -> str:
    if moneda == "USD":
        return _FORMATO_US_DOLARES
    # Una moneda que no es ni soles ni dólares no tiene símbolo contable
    # conocido: mejor su propio código que inventarle uno.
    return rf'_ * "{moneda}"\ #,##0.00_ ;_ * "{moneda}"\ \-#,##0.00_ ;_ * "-"??_ ;_ @_ '


class _Conversion(NamedTuple):
    """Cómo pasa a soles una fila del registro.

    `tipo_cambio is None` cubre dos casos que se tratan igual al escribir la
    fila: la moneda ya es soles, o es extranjera pero SUNAT no mandó tipo de
    cambio. La diferencia entre ambos la hace `factor`.
    """

    moneda: str
    tipo_cambio: float | None

    @property
    def factor(self) -> float | None:
        """`None` cuando hay moneda extranjera sin tipo de cambio: no hay con qué convertir."""
        if self.moneda == "PEN":
            return 1.0
        return self.tipo_cambio

    @property
    def formato_importe(self) -> str:
        return FORMATO_IMPORTE if self.factor is not None else _formato_sin_convertir(self.moneda)

    def soles(self, valor: Any, decimales: int = DECIMALES_IMPORTE) -> float | None:
        """El importe en soles, o el nominal si no hay tipo de cambio con qué convertir."""
        return _monto(valor, factor=self.factor or 1.0, decimales=decimales)

    def equivalente(self, total: Any) -> float | None:
        """Total en la moneda original, para la columna «EQUIVALENTE EN DOLARES»."""
        return None if self.moneda == "PEN" else _monto(total)


def _conversion(comprobante: dict[str, Any]) -> _Conversion:
    moneda = _texto(comprobante.get("moneda")) or "PEN"
    # `serializar` devuelve `0.0`, no `None`, cuando SUNAT no mandó tipo de
    # cambio (`monto_a_float`); `_monto` ya trata un cero como "sin valor".
    tipo_cambio = (
        None
        if moneda == "PEN"
        else _monto(comprobante.get("tipo_cambio"), decimales=DECIMALES_TIPO_CAMBIO)
    )
    return _Conversion(moneda=moneda, tipo_cambio=tipo_cambio)


def _analisis(comprobante: dict[str, Any]) -> dict[str, Any]:
    analisis = comprobante.get("analisis")
    return analisis if isinstance(analisis, dict) else {}


def _rag(comprobante: dict[str, Any]) -> dict[str, Any]:
    """Clasificación del RAG contable (`app/services/ollama_rag.py`), si corrió.

    Cuando existe, sus cuentas y su glosa son más precisas que las del
    análisis general —vienen de precedentes de la propia empresa, no sólo de
    una descripción— así que ganan a los campos planos del análisis.
    """
    rag = _analisis(comprobante).get("rag")
    return rag if isinstance(rag, dict) else {}


def _glosa(comprobante: dict[str, Any]) -> str | None:
    """Descripción corta de la operación para la columna de glosa, en mayúsculas.

    La glosa del RAG gana si existe. Si no, con un único ítem su nombre es la
    mejor glosa posible y casi siempre cabe entera; con varios, describir sólo
    el primero engañaría, así que se usa el resumen que hace la IA del
    comprobante completo. Siempre en mayúsculas, como la escriben los
    contadores en los cuatro registros reales revisados.
    """
    glosa_rag = _recortar(_rag(comprobante).get("glosa"), MAX_GLOSA)
    if glosa_rag:
        return glosa_rag.upper()

    analisis = _analisis(comprobante)
    detalle = analisis.get("detalle")
    if isinstance(detalle, list) and len(detalle) == 1 and isinstance(detalle[0], dict):
        glosa = _recortar(detalle[0].get("producto"), MAX_GLOSA)
        if glosa:
            return glosa.upper()
    glosa = _recortar(analisis.get("descripcion"), MAX_GLOSA)
    return glosa.upper() if glosa else None


def _cuenta_contable(comprobante: dict[str, Any]) -> str | None:
    cuenta = _rag(comprobante).get("cuenta_base") or _analisis(comprobante).get("cuenta_contable")
    return _recortar(cuenta, MAX_CUENTA_CONTABLE)


def _cuenta_total(comprobante: dict[str, Any], libro: Libro) -> str:
    """Cuenta contable del total: la que resolvió el RAG, o la general del libro."""
    return _texto(_rag(comprobante).get("cuenta_total")) or CUENTA_TOTAL[libro]


def _tasa_igv(comprobante: dict[str, Any]) -> float | int:
    """Tasa declarada por SUNAT; si no vino, la general. Siempre tiene valor.

    Los cuatro registros reales con los que se comparó esta exportación
    llevan una tasa en todas las filas, también en las operaciones sin IGV.
    """
    return _monto(comprobante.get("porcentaje_igv")) or TASA_IGV


def _condicion_pago(comprobante: dict[str, Any]) -> str:
    """CRE cuando hay un vencimiento posterior a la emisión; si no, CON.

    Ninguna fuente de SUNAT manda la condición de pago: es una heurística a
    partir de las fechas, la misma señal que usa un contador. Acertó en torno
    al 92 % de las compras de uno de los clientes usados para compararla.
    """
    emision = normalizar_fecha(comprobante.get("fecha_emision"))
    vencimiento = normalizar_fecha(comprobante.get("fecha_vencimiento"))
    if emision is not None and vencimiento is not None and vencimiento > emision:
        return CONDICION_CREDITO
    return CONDICION_CONTADO


def _primero(valores: dict[str, Any], claves: tuple[str, ...]) -> Any:
    for clave in claves:
        valor = valores.get(clave)
        if valor not in (None, ""):
            return valor
    return None


# Candidatos para los datos del comprobante que modifica una nota de crédito o
# débito. El RVIE los manda dentro de `documentoMod`; ningún fixture del
# repositorio trae la fecha, así que se deja un segundo nombre candidato.
_CAMPOS_MOD_FECHA = ("fecEmision", "fecEmisionMod")
_CAMPOS_MOD_TIPO = ("codTipoCDP",)
_CAMPOS_MOD_SERIE = ("numSerieCDP",)
_CAMPOS_MOD_NUMERO = ("numCDP",)


def _referencia_modificado(comprobante: dict[str, Any]) -> dict[str, Any]:
    """Fecha/tipo/serie/número del comprobante que modifica una NC o ND, en ventas.

    Sólo el RVIE manda `documentoMod` (el RCE no lo hace nunca, así que en
    compras estas columnas se quedan vacías). No se filtra por `tipo_cp`
    07/08: SUNAT sólo llena esta lista a quien la tiene.
    """
    modificados = comprobante.get("documentos_modificados")
    if not isinstance(modificados, list) or not modificados:
        return {}
    original = next((item for item in modificados if isinstance(item, dict)), None)
    if original is None:
        return {}
    return {
        "R": normalizar_fecha(_primero(original, _CAMPOS_MOD_FECHA)),
        "S": normalizar_tipo_cp(_primero(original, _CAMPOS_MOD_TIPO)) or None,
        "T": _texto(_primero(original, _CAMPOS_MOD_SERIE)),
        "U": _entero_o_texto(_primero(original, _CAMPOS_MOD_NUMERO)),
    }


def _fechas(comprobante: dict[str, Any]) -> tuple[date | None, date | None]:
    emision = normalizar_fecha(comprobante.get("fecha_emision"))
    vencimiento = normalizar_fecha(comprobante.get("fecha_vencimiento")) or emision
    return emision, vencimiento


def _fila_compras(
    comprobante: dict[str, Any],
    conversion: _Conversion,
    destino: str | None = None,
) -> dict[str, Any]:
    emision, vencimiento = _fechas(comprobante)
    rag = _rag(comprobante)

    destino_efectivo = (destino or comprobante.get("destino_compras") or "").lower()
    base_gravada = comprobante.get("base_imponible_dg") or comprobante.get("base_imponible")
    igv_gravado = comprobante.get("igv_dg") or comprobante.get("igv")

    if destino_efectivo == "dng":
        # Destinado a operaciones no gravadas (empresas con ventas exoneradas o compras a costo)
        col_j, col_k = None, None
        col_l, col_m = None, None
        col_n = conversion.soles(base_gravada)
        col_o = conversion.soles(igv_gravado)
    elif destino_efectivo == "dgng":
        # Destinado a operaciones gravadas y no gravadas (prorrata)
        col_j, col_k = None, None
        col_l = conversion.soles(base_gravada)
        col_m = conversion.soles(igv_gravado)
        col_n, col_o = None, None
    elif destino_efectivo == "dg":
        # Destinado a operaciones gravadas (crédito fiscal pleno)
        col_j = conversion.soles(base_gravada)
        col_k = conversion.soles(igv_gravado)
        col_l, col_m = None, None
        col_n, col_o = None, None
    else:
        # Desglose por defecto respetando los campos desagregados de SUNAT SIRE
        col_j = conversion.soles(comprobante.get("base_imponible_dg"))
        col_k = conversion.soles(comprobante.get("igv_dg"))
        col_l = conversion.soles(comprobante.get("base_imponible_dgng"))
        col_m = conversion.soles(comprobante.get("igv_dgng"))
        col_n = conversion.soles(comprobante.get("base_imponible_dng"))
        col_o = conversion.soles(comprobante.get("igv_dng"))

    return {
        "A": emision,
        "B": vencimiento,
        "C": _texto(rag.get("codigo_comprobante") or comprobante.get("tipo_cp")),
        "D": _texto(comprobante.get("serie")),
        "F": _entero_o_texto(comprobante.get("numero")),
        "G": _entero_o_texto(
            rag.get("codigo_identidad") or comprobante.get("tipo_doc_identidad")
        ),
        "H": _entero_o_texto(comprobante.get("documento_contraparte")),
        "I": _texto(comprobante.get("razon_social")),
        "J": col_j,
        "K": col_k,
        "L": col_l,
        "M": col_m,
        "N": col_n,
        "O": col_o,
        # Adquisiciones no gravadas. El SIRE las manda agrupadas en
        # `no_gravado`; sin sumarlo esta columna salía en cero para todas
        # las compras exoneradas o inafectas.
        "P": conversion.soles(
            _suma(
                comprobante.get("exonerado"),
                comprobante.get("inafecto"),
                comprobante.get("no_gravado"),
            )
        ),
        "Q": conversion.soles(comprobante.get("isc")),
        "R": conversion.soles(comprobante.get("otros_tributos")),
        "S": conversion.soles(comprobante.get("total")),
        "W": conversion.tipo_cambio,
        "AB": _moneda(comprobante.get("moneda")),
        "AC": conversion.equivalente(comprobante.get("total")),
        "AD": vencimiento,
        "AE": _condicion_pago(comprobante),
        "AF": _cuenta_contable(comprobante),
        "AH": _cuenta_total(comprobante, Libro.COMPRAS),
        "AR": _tasa_igv(comprobante),
        "AS": _glosa(comprobante),
    }


def _fila_ventas(comprobante: dict[str, Any], conversion: _Conversion) -> dict[str, Any]:
    emision, vencimiento = _fechas(comprobante)
    rag = _rag(comprobante)
    return {
        "A": emision,
        "B": vencimiento,
        "C": _texto(rag.get("codigo_comprobante") or comprobante.get("tipo_cp")),
        "D": _texto(comprobante.get("serie")),
        "E": _entero_o_texto(comprobante.get("numero")),
        "F": _entero_o_texto(
            rag.get("codigo_identidad") or comprobante.get("tipo_doc_identidad")
        ),
        "G": _entero_o_texto(comprobante.get("documento_contraparte")),
        "H": _texto(comprobante.get("razon_social")),
        "J": conversion.soles(comprobante.get("base_imponible")),
        "K": conversion.soles(comprobante.get("exonerado")),
        "L": conversion.soles(comprobante.get("inafecto")),
        "M": conversion.soles(comprobante.get("isc")),
        "N": conversion.soles(comprobante.get("igv")),
        "O": conversion.soles(comprobante.get("otros_tributos")),
        "P": conversion.soles(comprobante.get("total")),
        "Q": conversion.tipo_cambio,
        "V": _moneda(comprobante.get("moneda")),
        "W": conversion.equivalente(comprobante.get("total")),
        "X": vencimiento,
        "Y": _condicion_pago(comprobante),
        # En la hoja de ventas la moneda es la V: AB es la cuenta contable.
        "AB": _cuenta_contable(comprobante),
        "AD": _cuenta_total(comprobante, Libro.VENTAS),
        "AL": _tasa_igv(comprobante),
        "AM": _glosa(comprobante),
        **_referencia_modificado(comprobante),
    }


# Columnas de fecha y celda que rotula el pie de totales. `R` es la fecha del
# comprobante que modifica una nota de crédito o débito, en ventas.
_COLUMNAS_FECHA = {Libro.COMPRAS: ("A", "B", "AD"), Libro.VENTAS: ("A", "B", "R", "X")}
_ROTULO_TOTAL = {Libro.COMPRAS: "A", Libro.VENTAS: "G"}

# Las columnas de dinero: son las que llevan el formato contable y las que
# suma el pie. Coinciden porque todo lo que es un importe se totaliza. El
# tipo de cambio, el equivalente en dólares y el porcentaje de IGV quedan
# fuera a propósito: no son importes en soles del propio comprobante.
_COLUMNAS_IMPORTE = {
    Libro.COMPRAS: ("J", "K", "L", "M", "N", "O", "P", "Q", "R", "S"),
    Libro.VENTAS: ("J", "K", "L", "M", "N", "O", "P"),
}

_MAPEOS = {Libro.COMPRAS: _fila_compras, Libro.VENTAS: _fila_ventas}


def _prototipos(hoja: Worksheet) -> dict[str, Any]:
    """Estilo de cada columna en la primera fila de datos de la plantilla."""
    return {celda.column_letter: celda._style for celda in hoja[PRIMERA_FILA_DATOS]}


def _limpiar_ejemplos(hoja: Worksheet) -> None:
    """Borra las filas de ejemplo y el pie de totales que trae la plantilla.

    La plantilla actual no trae filas de ejemplo por debajo de la de
    prototipo, pero se mantiene el borrado por si una plantilla futura sí las
    trae. Los rangos combinados del encabezado están en las filas 1-3, así
    que sólo hay que desarmar los del pie antes de borrar.
    """
    for rango in list(hoja.merged_cells.ranges):
        if rango.min_row >= PRIMERA_FILA_DATOS:
            hoja.unmerge_cells(str(rango))

    if hoja.max_row >= PRIMERA_FILA_DATOS:
        hoja.delete_rows(PRIMERA_FILA_DATOS, hoja.max_row - PRIMERA_FILA_DATOS + 1)

    # `delete_rows` no desplaza `row_dimensions`: sin esto, una plantilla con
    # alturas de fila explícitas por debajo de la de prototipo las dejaría
    # ahí, mezclándose con las filas nuevas.
    for fila in [f for f in hoja.row_dimensions if f >= PRIMERA_FILA_DATOS]:
        del hoja.row_dimensions[fila]


def _escribir_fila(
    hoja: Worksheet,
    indice: int,
    valores: dict[str, Any],
    prototipos: dict[str, Any],
    columnas_fecha: tuple[str, ...],
    columnas_importe: tuple[str, ...],
    formato_importe: str,
) -> None:
    for letra, estilo in prototipos.items():
        hoja[f"{letra}{indice}"]._style = copy(estilo)

    for letra, valor in valores.items():
        if valor is None:
            continue
        celda = hoja[f"{letra}{indice}"]
        celda.value = valor
        if letra in columnas_fecha:
            celda.number_format = FORMATO_FECHA
        elif letra in columnas_importe:
            # El símbolo (si lo hay) va por fila, no por columna: en un mismo
            # registro conviven comprobantes que sí se pudieron convertir a
            # soles con alguno que no.
            celda.number_format = formato_importe


def _escribir_totales(hoja: Worksheet, indice: int, ultima: int, libro: Libro) -> None:
    """Un único pie de totales: todo el registro ya está en soles."""
    hoja[f"{_ROTULO_TOTAL[libro]}{indice}"] = "TOTAL"
    for letra in _COLUMNAS_IMPORTE[libro]:
        celda = hoja[f"{letra}{indice}"]
        celda.value = f"=SUM({letra}{PRIMERA_FILA_DATOS}:{letra}{ultima})"
        celda.number_format = FORMATO_IMPORTE


def excel_plantilla(
    comprobantes: list[dict[str, Any]],
    libro: Libro,
    destino: str | None = None,
) -> io.BytesIO:
    """Genera el registro del libro pedido sobre la plantilla oficial."""
    wb = load_workbook(io.BytesIO(_bytes_plantilla()))

    for nombre in wb.sheetnames:
        if nombre != HOJAS[libro]:
            wb.remove(wb[nombre])

    hoja = wb[HOJAS[libro]]
    hoja.sheet_view.topLeftCell = f"A{PRIMERA_FILA_DATOS}"
    if hoja.sheet_view.pane:
        hoja.sheet_view.pane.topLeftCell = f"A{PRIMERA_FILA_DATOS}"

    prototipos = _prototipos(hoja)
    _limpiar_ejemplos(hoja)

    columnas_fecha = _COLUMNAS_FECHA[libro]
    columnas_importe = _COLUMNAS_IMPORTE[libro]

    indice = PRIMERA_FILA_DATOS
    for comprobante in comprobantes:
        conversion = _conversion(comprobante)
        valores = (
            _fila_compras(comprobante, conversion, destino=destino)
            if libro == Libro.COMPRAS
            else _fila_ventas(comprobante, conversion)
        )
        _escribir_fila(
            hoja,
            indice,
            valores,
            prototipos,
            columnas_fecha,
            columnas_importe,
            conversion.formato_importe,
        )
        indice += 1

    if comprobantes:
        _escribir_totales(hoja, indice, indice - 1, libro)

    salida = io.BytesIO()
    wb.save(salida)
    salida.seek(0)
    return salida
