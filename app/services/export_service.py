from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def _texto(valor: Any) -> str:
    if isinstance(valor, (dict, list)):
        return str(valor)
    return "" if valor is None else str(valor)


def _resultado(valor: Any) -> str:
    texto = _texto(valor).strip().upper()
    return "NO DETERMINADO" if texto in {"", "NODETERMINADO"} else texto


def _money(valor: Any) -> str:
    try:
        return f"{float(valor):.2f}"
    except (TypeError, ValueError):
        return _texto(valor)


def _analisis(comprobante: dict[str, Any]) -> dict[str, Any]:
    return comprobante.get("analisis") or {}


def _consistencia(comprobante: dict[str, Any]) -> str:
    analisis = _analisis(comprobante)
    detalle = analisis.get("detalle") if isinstance(analisis.get("detalle"), list) else []

    try:
        total = float(comprobante.get("total"))
    except (TypeError, ValueError):
        total = None

    if not detalle:
        return "Detalle inferido"
    if total is None:
        return "Revision recomendada"

    suma = 0.0
    for item in detalle:
        if not isinstance(item, dict):
            continue
        try:
            suma += float(item.get("importe", 0) or 0)
        except (TypeError, ValueError):
            continue

    if abs(suma - total) <= 0.01:
        return "Detalle completo"
    if 0 < suma < total:
        return "Detalle inferido"
    return "Revision recomendada"


def excel_de_comprobante(comprobante: dict[str, Any]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comprobante"

    ws.append(["Campo", "Valor"])
    negrita = Font(bold=True)
    for celda in ws[1]:
        celda.font = negrita
        celda.alignment = Alignment(horizontal="center")

    for clave, valor in comprobante.items():
        ws.append([str(clave), _texto(valor)])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50

    salida = io.BytesIO()
    wb.save(salida)
    salida.seek(0)
    return salida


def pdf_de_comprobante(comprobante: dict[str, Any]) -> io.BytesIO:
    salida = io.BytesIO()
    doc = SimpleDocTemplate(
        salida, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40
    )
    elementos = []
    estilos = getSampleStyleSheet()

    titulo = ParagraphStyle(
        "Titulo", parent=estilos["Heading1"], fontName="Helvetica-Bold",
        fontSize=16, textColor=colors.HexColor("#333333"), spaceAfter=2,
    )
    subtitulo = ParagraphStyle(
        "Subtitulo", parent=estilos["Normal"], fontName="Helvetica",
        fontSize=10, textColor=colors.HexColor("#666666"), spaceAfter=20,
    )
    seccion = ParagraphStyle(
        "Seccion", parent=estilos["Heading2"], fontName="Helvetica-Bold",
        fontSize=11, textColor=colors.HexColor("#1976d2"), spaceBefore=15, spaceAfter=8,
    )
    normal = estilos["Normal"]
    cursiva = ParagraphStyle(
        "Cursiva", parent=estilos["Normal"], fontName="Helvetica-Oblique",
        textColor=colors.HexColor("#555555"),
    )

    analisis = _analisis(comprobante)
    moneda = comprobante.get("moneda", "PEN")
    total = _money(comprobante.get("total", 0))

    elementos.append(Paragraph(_texto(comprobante.get("serie_numero")) or "Comprobante", titulo))
    elementos.append(Paragraph(_texto(comprobante.get("razon_social")) or "Contraparte", subtitulo))

    elementos.append(Paragraph("INFORMACIÓN GENERAL", seccion))
    tabla_gen = Table(
        [[
            Paragraph("<b>Tipo:</b>", normal),
            Paragraph(_texto(comprobante.get("tipo_cp_descripcion")), normal),
            Paragraph("<b>Fecha de Emisión:</b>", normal),
            Paragraph(_texto(comprobante.get("fecha_emision")) or "-", normal),
        ]],
        colWidths=[60, 200, 110, 110],
    )
    tabla_gen.setStyle(
        TableStyle([("ALIGN", (0, 0), (-1, -1), "LEFT"), ("VALIGN", (0, 0), (-1, -1), "TOP")])
    )
    elementos.append(tabla_gen)

    elementos.append(Paragraph("CLASIFICACIÓN CONTABLE", seccion))
    tabla_clasif = Table(
        [[
            Paragraph(f"<b>Resultado:</b> {_resultado(analisis.get('resultado'))}", normal),
            Paragraph(f"<b>Confianza:</b> {_texto(analisis.get('confianza')) or '0%'}", normal),
        ]],
        colWidths=[200, 300],
    )
    tabla_clasif.setStyle(
        TableStyle([("ALIGN", (0, 0), (-1, -1), "LEFT"), ("VALIGN", (0, 0), (-1, -1), "TOP")])
    )
    elementos.append(tabla_clasif)
    elementos.append(Spacer(1, 4))
    observaciones = _texto(analisis.get("observaciones")) or "Sin observaciones detalladas."
    elementos.append(Paragraph(f"<i>{observaciones}</i>", cursiva))

    documento = _texto(comprobante.get("documento_contraparte")) or "-"
    razon_social = _texto(comprobante.get("razon_social")) or "-"
    serie_numero = _texto(comprobante.get("serie_numero"))
    estado = _texto(comprobante.get("estado_procesamiento"))

    tabla_partes = Table(
        [
            [
                Paragraph("DATOS DE LA CONTRAPARTE", seccion),
                Paragraph("DATOS DEL COMPROBANTE", seccion),
            ],
            [
                Paragraph(f"<b>Documento:</b> {documento}", normal),
                Paragraph(f"<b>Serie-Número:</b> {serie_numero}", normal),
            ],
            [
                Paragraph(f"<b>Razón social:</b> {razon_social}", normal),
                Paragraph(f"<b>Moneda:</b> {moneda}", normal),
            ],
            ["", Paragraph(f"<b>Estado:</b> {estado}", normal)],
        ],
        colWidths=[260, 260],
    )
    tabla_partes.setStyle(
        TableStyle([("ALIGN", (0, 0), (-1, -1), "LEFT"), ("VALIGN", (0, 0), (-1, -1), "TOP")])
    )
    elementos.append(tabla_partes)

    base = _money(comprobante.get("base_imponible"))

    elementos.append(Paragraph("MONTOS", seccion))
    tabla_montos = Table(
        [[
            Paragraph(f"<b>Base imponible:</b> {base}", normal),
            Paragraph(f"<b>IGV:</b> {_money(comprobante.get('igv'))}", normal),
            Paragraph(f"<b>Total:</b> {moneda} {total}", normal),
        ]],
        colWidths=[173, 173, 173],
    )
    tabla_montos.setStyle(
        TableStyle([("ALIGN", (0, 0), (-1, -1), "LEFT"), ("VALIGN", (0, 0), (-1, -1), "TOP")])
    )
    elementos.append(tabla_montos)

    elementos.append(Paragraph("RESUMEN", seccion))
    detalle = analisis.get("detalle")

    if detalle and isinstance(detalle, list):
        for idx, item in enumerate(detalle):
            if not isinstance(item, dict):
                continue
            producto = _texto(item.get("producto")) or "Item general / No especificado"
            categoria = _texto(item.get("categoria_contable")) or "-"
            cantidad = _texto(item.get("cantidad")) or "N/A"
            importe = _money(item.get("importe", 0))
            razon = _texto(item.get("razon"))

            filas = [
                [Paragraph(f"<b>{producto}</b>", normal)],
                [
                    Paragraph(
                        f"{categoria} — Cant: {cantidad} — Importe: {moneda} {importe}", normal
                    )
                ],
            ]
            if razon:
                filas.append([Paragraph(f"<i>{razon}</i>", cursiva)])

            tabla_item = Table(filas, colWidths=[520])
            tabla_item.setStyle(
                TableStyle([
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("BOTTOMPADDING", (0, -1), (-1, -1), 8),
                ])
            )
            elementos.append(tabla_item)
            if idx < len(detalle) - 1:
                elementos.append(Spacer(1, 4))
    else:
        elementos.append(Paragraph("No hay detalle analizado.", normal))

    doc.build(elementos)
    salida.seek(0)
    return salida


def excel_de_lote(comprobantes: list[dict[str, Any]]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comprobantes"

    ws.append([
        "SERIE_NUMERO",
        "TIPO",
        "DOCUMENTO",
        "RAZON_SOCIAL",
        "FECHA_EMISION",
        "BASE_IMPONIBLE",
        "IGV",
        "TOTAL",
        "RESULTADO",
        "CONSISTENCIA",
        "ESTADO",
    ])

    negrita = Font(bold=True)
    for celda in ws[1]:
        celda.font = negrita
        celda.alignment = Alignment(horizontal="center")

    for comprobante in comprobantes:
        analisis = _analisis(comprobante)
        ws.append([
            _texto(comprobante.get("serie_numero")),
            _texto(comprobante.get("tipo_cp_descripcion")),
            _texto(comprobante.get("documento_contraparte")),
            _texto(comprobante.get("razon_social")),
            _texto(comprobante.get("fecha_emision")),
            _texto(comprobante.get("base_imponible")),
            _texto(comprobante.get("igv")),
            _texto(comprobante.get("total")),
            _resultado(analisis.get("resultado")),
            _consistencia(comprobante),
            _texto(comprobante.get("estado_procesamiento")),
        ])

    anchos = {
        "A": 20, "B": 30, "C": 16, "D": 38, "E": 16, "F": 16,
        "G": 14, "H": 14, "I": 18, "J": 22, "K": 18,
    }
    for columna, ancho in anchos.items():
        ws.column_dimensions[columna].width = ancho

    salida = io.BytesIO()
    wb.save(salida)
    salida.seek(0)
    return salida


def pdf_de_lote(comprobantes: list[dict[str, Any]]) -> io.BytesIO:
    salida = io.BytesIO()
    doc = SimpleDocTemplate(
        salida, pagesize=letter, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36
    )
    elementos = []
    estilos = getSampleStyleSheet()

    titulo = ParagraphStyle(
        "Titulo", parent=estilos["Title"], fontName="Helvetica-Bold",
        fontSize=18, textColor=colors.HexColor("#1e293b"),
    )
    subtitulo = ParagraphStyle(
        "Subtitulo", parent=estilos["Normal"], fontSize=11,
        textColor=colors.HexColor("#475569"), spaceAfter=20,
    )
    item_titulo = ParagraphStyle(
        "ItemTitulo", parent=estilos["Heading2"], fontName="Helvetica-Bold",
        fontSize=11, textColor=colors.HexColor("#0f172a"), spaceBefore=12, spaceAfter=2,
    )
    item_cuerpo = ParagraphStyle(
        "ItemCuerpo", parent=estilos["Normal"], fontName="Helvetica",
        fontSize=9, leading=14, spaceAfter=2,
    )
    item_obs = ParagraphStyle(
        "ItemObs", parent=estilos["Normal"], fontName="Helvetica-Oblique",
        fontSize=9, leading=12, textColor=colors.HexColor("#334155"), spaceAfter=8,
    )

    elementos.append(Paragraph("Reporte Detallado de Comprobantes", titulo))

    acumulado = 0.0
    for comprobante in comprobantes:
        try:
            acumulado += float(comprobante.get("total") or 0)
        except (TypeError, ValueError):
            continue

    elementos.append(
        Paragraph(
            f"Total comprobantes: {len(comprobantes)} &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"Monto acumulado: {acumulado:,.2f}",
            subtitulo,
        )
    )

    for comprobante in comprobantes[:500]:
        analisis = _analisis(comprobante)
        referencia = _texto(comprobante.get("serie_numero"))
        contraparte = _texto(comprobante.get("razon_social")) or "Contraparte desconocida"
        observaciones = (
            _texto(analisis.get("observaciones")) or "Sin observaciones detalladas registradas."
        )

        elementos.append(Paragraph(f"<b>{referencia}</b> — {contraparte}", item_titulo))
        moneda = comprobante.get("moneda", "PEN")
        total = _money(comprobante.get("total"))
        resultado = _resultado(analisis.get("resultado"))
        confianza = _texto(analisis.get("confianza")) or "0%"
        separador_campos = " &nbsp;&nbsp;|&nbsp;&nbsp; "

        elementos.append(
            Paragraph(
                f"<b>Total:</b> {moneda} {total}{separador_campos}"
                f"<b>Resultado:</b> {resultado}{separador_campos}"
                f"<b>Confianza IA:</b> {confianza}",
                item_cuerpo,
            )
        )
        elementos.append(Paragraph(f"<i>Justificación: {observaciones}</i>", item_obs))

        separador = Table([[""]], colWidths=["100%"])
        separador.setStyle(
            TableStyle([
                ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ])
        )
        elementos.append(separador)

    doc.build(elementos)
    salida.seek(0)
    return salida
