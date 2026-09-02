"""Indexa los Excel RAG en colecciones Mongo separadas usando Ollama.

Uso: uv run python scripts/indexar_rag_contable.py [--rehacer] [--dry-run]
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import hashlib
import sys
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from langchain_core.documents import Document
from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import load_workbook

RAIZ = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(RAIZ))
load_dotenv(RAIZ / ".env")

from app.core.config import settings  # noqa: E402
from app.repositories._mongo import (  # noqa: E402
    NOMBRE_COL_RAG_ACCOUNTS,
    NOMBRE_COL_RAG_COMPANY,
    NOMBRE_COL_RAG_HISTORICAL,
    NOMBRE_COL_RAG_RULES,
    NOMBRE_COL_RAG_TAX,
)
from app.services.ollama_rag import embed_documents  # noqa: E402

CARPETA = RAIZ / "source" / "rag"
DATOS_CONTASIS = RAIZ / "app" / "resources" / "rag_contasis"
RUC_POR_EMPRESA = {"JOAQUISAN": "20608997106"}
IGNORADAS = {"00_LEEME", "04_CAPTURA_SIRE", "05_AUDITORIA_RAG", "06_CATALOGOS", "07_VERSIONADO"}


def _valor(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _coleccion(archivo: str, hoja: str) -> tuple[str, str] | None:
    if archivo.startswith("RAG_01"):
        if hoja == "05_CUENTAS_CONTASIS":
            return NOMBRE_COL_RAG_ACCOUNTS, "account"
        if hoja == "06_COMPROBANTES":
            return NOMBRE_COL_RAG_TAX, "tax"
        return NOMBRE_COL_RAG_RULES, "rule"
    if archivo.startswith("RAG_02"):
        return NOMBRE_COL_RAG_HISTORICAL, "historical"
    if archivo.startswith("RAG_03") and hoja in {
        "01_PERFIL_EMPRESA",
        "02_OVERRIDES",
        "03_REFERENCIAS",
    }:
        return NOMBRE_COL_RAG_COMPANY, "company"
    return None


def cargar() -> list[tuple[str, Document]]:
    salida = []
    for ruta in sorted(CARPETA.glob("RAG_*.xlsx")):
        wb = load_workbook(ruta, read_only=True, data_only=True)
        for ws in wb.worksheets:
            if ws.title in IGNORADAS:
                continue
            destino = _coleccion(ruta.name, ws.title)
            if not destino:
                continue
            coleccion, source_type = destino
            filas = ws.iter_rows(values_only=True)
            encabezados = [_valor(v) for v in next(filas, [])]
            for numero, valores in enumerate(filas, start=2):
                datos = {k: _valor(v) for k, v in zip(encabezados, valores, strict=False) if k}
                if not any(datos.values()):
                    continue
                texto = ". ".join(f"{k}: {v}" for k, v in datos.items() if v)
                row_id = (
                    datos.get("rule_id")
                    or datos.get("example_id")
                    or datos.get("override_id")
                    or datos.get("reference_id")
                    or datos.get("codigo")
                    or f"{ws.title}:{numero}"
                )
                metadata = {
                    "source_type": source_type,
                    "file": ruta.name,
                    "sheet": ws.title,
                    "row_id": row_id,
                    "empresa": datos.get("empresa", ""),
                    "registro": datos.get("registro", ""),
                    "clasificacion": datos.get("clasificacion")
                    or datos.get("clasificacion_esperada", ""),
                    "code": datos.get("codigo") or datos.get("cuenta_base", ""),
                    "description": datos.get("descripcion") or datos.get("cuenta_base_desc", ""),
                }
                salida.append((coleccion, Document(page_content=texto, metadata=metadata)))
        wb.close()
    fuentes_csv = (
        ("plan_cuentas.csv", NOMBRE_COL_RAG_ACCOUNTS, "account"),
        ("historicos.csv", NOMBRE_COL_RAG_HISTORICAL, "historical"),
    )
    for archivo, coleccion, source_type in fuentes_csv:
        ruta = DATOS_CONTASIS / archivo
        with ruta.open(encoding="utf-8-sig", newline="") as stream:
            for numero, datos in enumerate(csv.DictReader(stream), start=2):
                texto = ". ".join(f"{k}: {_valor(v)}" for k, v in datos.items() if _valor(v))
                empresa = _valor(datos.get("empresa"))
                row_id = _valor(datos.get("codigo")) or f"{archivo}:{numero}"
                metadata = {
                    "source_type": source_type,
                    "file": archivo,
                    "sheet": _valor(datos.get("source_sheet")),
                    "row_id": row_id,
                    "empresa": empresa,
                    "empresa_ruc": RUC_POR_EMPRESA.get(empresa, ""),
                    "registro": _valor(datos.get("registro")),
                    "code": _valor(datos.get("codigo") or datos.get("cuenta_base")),
                    "description": _valor(
                        datos.get("descripcion") or datos.get("cuenta_base_desc")
                    ),
                    "cuenta_total": _valor(datos.get("cuenta_total")),
                    "glosa": _valor(datos.get("glosa")),
                    "proveedor_ruc": _valor(datos.get("numero_identidad")),
                }
                salida.append((coleccion, Document(page_content=texto, metadata=metadata)))
    return salida


async def ejecutar(rehacer: bool, dry_run: bool) -> None:
    docs = cargar()
    print(f"Filas válidas: {len(docs)}")
    if dry_run:
        return
    cliente = AsyncIOMotorClient(settings.MONGO_URI)
    db = cliente[settings.MONGO_FACTURASDB_NAME]
    colecciones = sorted({c for c, _ in docs})
    if rehacer:
        for nombre in colecciones:
            await db[nombre].delete_many({})
    for nombre in colecciones:
        grupo = [doc for coleccion, doc in docs if coleccion == nombre]
        for inicio in range(0, len(grupo), 16):
            lote = grupo[inicio : inicio + 16]
            claves = [
                hashlib.sha256(
                    f"{doc.metadata['file']}|{doc.metadata['sheet']}|"
                    f"{doc.metadata['row_id']}|{doc.page_content}".encode()
                ).hexdigest()
                for doc in lote
            ]
            existentes = {
                fila["content_hash"]
                async for fila in db[nombre].find(
                    {"content_hash": {"$in": claves}}, {"content_hash": 1}
                )
            }
            pendientes = [
                (doc, clave)
                for doc, clave in zip(lote, claves, strict=True)
                if clave not in existentes
            ]
            if not pendientes:
                continue
            lote = [doc for doc, _ in pendientes]
            # Los precedentes se resuelven por empresa, proveedor y glosa antes
            # de cualquier similitud. Evitamos vectorizar miles de históricos
            # para que la reconstrucción sea viable en CPU.
            if nombre == NOMBRE_COL_RAG_HISTORICAL:
                vectores = [[] for _ in lote]
            else:
                vectores = await asyncio.to_thread(embed_documents, [d.page_content for d in lote])
            for (doc, clave), vector in zip(pendientes, vectores, strict=True):
                await db[nombre].replace_one(
                    {"content_hash": clave},
                    {
                        "content_hash": clave,
                        "texto": doc.page_content,
                        "metadata": doc.metadata,
                        "embedding": vector,
                    },
                    upsert=True,
                )
        await db[nombre].create_index("content_hash", unique=True)
        await db[nombre].create_index([("metadata.empresa", 1), ("metadata.registro", 1)])
        print(f"{nombre}: {len(grupo)}")
    cliente.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--rehacer", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    asyncio.run(ejecutar(args.rehacer, args.dry_run))
