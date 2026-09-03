"""Recorta y agranda la plantilla oficial «REGISTRO DE COMPRAS Y VENTAS» de Contasis.

`app/resources/plantilla_registro.xlsx` es la que usa [plantilla_excel.py]
(../app/services/plantilla_excel.py) para exportar; venía con letra diminuta,
trece filas de notas y especificación antes del encabezado, y columnas tan
angostas que el propio encabezado se cortaba. Los registros reales que llenan
los contadores (`source/REGISTROS CASOS REALES/`) no traen nada de eso: el
encabezado de tres niveles ocupa las filas 1-3 y los datos empiezan en la 4.

Este script hace la misma poda sobre la plantilla original de Contasis
(`source/PLANTILLA REGISTRO DE COMPRAS Y VENTAS.xlsx`, fuera del control de
versiones) y escribe el resultado en `app/resources/plantilla_registro.xlsx`:

- Borra las notas de uso, el título y la fila de especificacion (filas 1-13
  originales); el encabezado de tres niveles queda en las filas 1-3.
- Borra las filas de ejemplo y el pie que trae la plantilla; sólo sobrevive la
  fila 4, vacía, con el estilo que `plantilla_excel.py` usa como prototipo.
- Repinta el encabezado: negrita, blanco sobre azul, bordes finos, como los
  registros reales.
- Ensancha cada columna a partir de la palabra más larga de su encabezado
  (nunca por debajo de un mínimo legible) sin mover ninguna columna de sitio,
  porque Contasis importa por posición.
- Dos columnas de ventas que la plantilla trae ocultas (INAFECTA, ISC) las
  llena `plantilla_excel.py`; se muestran para que no queden datos invisibles.
- Inmoviliza los paneles en A4 y quita el autofiltro y los formatos
  condicionales que apuntaban a filas que ya no existen.

No toca `app/services/plantilla_excel.py`: sólo prepara el archivo que ese
módulo consume. Se ejecuta una vez y el resultado se versiona en git.

    uv run python scripts/preparar_plantilla.py
    uv run python scripts/preparar_plantilla.py \
        --origen otra_plantilla.xlsx --destino /tmp/prueba.xlsx
"""

from __future__ import annotations

import argparse
import re
import zipfile
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.worksheet.worksheet import Worksheet

REPO = Path(__file__).resolve().parents[1]
ORIGEN_DEFECTO = REPO / "source" / "PLANTILLA REGISTRO DE COMPRAS Y VENTAS.xlsx"
DESTINO_DEFECTO = REPO / "app" / "resources" / "plantilla_registro.xlsx"

# En la plantilla original el encabezado de tres niveles vive en las filas
# 10-12, la 13 es la fila azul de especificación ("dd/mm/yyyy", "02
# CARACTERES"...) y la 14 es la primera fila de ejemplo. Todo lo de encima
# -notas de uso y título- desaparece junto con la especificación.
FILA_HEADER_INICIO, FILA_HEADER_FIN = 10, 12
FILA_SPEC = 13
FILA_PROTOTIPO = 14
DESPLAZAMIENTO = FILA_HEADER_INICIO - 1  # 9: cuánto sube el encabezado.
ULTIMA_COL_TAIL = 16384  # Fin de la hoja: el grupo de columnas "de cola".

# Columnas que la plantilla trae ocultas pero que sí llena la exportación
# (ver `_fila_ventas` en plantilla_excel.py): quedarían con datos invisibles.
COLUMNAS_A_MOSTRAR = {"FORMATO_VENTAS": ("L", "M", "O")}

AZUL = "0070C0"
FUENTE_HEADER = {"name": "Calibri", "size": 9, "bold": True, "color": "FFFFFF"}
FUENTE_DATOS = ("Calibri", 10.0)
# 32/32 en vez de 24/24: con las columnas ensanchadas el encabezado más largo
# (compras "ADQUISICIONES GRAVADAS DESTINADAS A OPERACIONES GRAVADAS Y/O DE
# EXPORTACION Y A OPERACIONES NO GRAVADAS") todavía necesita varias líneas.
ALTURAS_HEADER = {1: 32.0, 2: 32.0, 3: 60.0}
ALTURA_DATOS = 18.75
ANCHO_MINIMO = 6.0
RELLENO_PALABRA = 2.0
PUNTOS_POR_LINEA = 12.0  # Aproximado para Calibri 9 en negrita.

_LADO = Side(style="thin")
BORDE = Border(left=_LADO, right=_LADO, top=_LADO, bottom=_LADO)
RELLENO = PatternFill("solid", fgColor=AZUL)
ALINEACION = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _rangos_header(ws: Worksheet) -> list[CellRange]:
    """Combinaciones del encabezado (filas 10-12), sin el título ni el pie."""
    return [
        CellRange(str(r))
        for r in ws.merged_cells.ranges
        if r.min_row >= FILA_HEADER_INICIO and r.max_row <= FILA_HEADER_FIN
    ]


def _ultima_columna_header(ws: Worksheet) -> int:
    ultima = 0
    for fila in range(FILA_HEADER_INICIO, FILA_HEADER_FIN + 1):
        for celda in ws[fila]:
            if celda.value is not None:
                ultima = max(ultima, celda.column)
    return ultima


def _recortar_filas(ws: Worksheet) -> list[CellRange]:
    """Desarma el encabezado, borra lo que sobra y vuelve a armarlo 9 filas arriba.

    `delete_rows` no desplaza `merged_cells`: hay que desmerger antes de borrar
    y volver a mergear con los rangos ya trasladados, o el archivo termina con
    combinaciones que apuntan a filas equivocadas.
    """
    rangos = _rangos_header(ws)
    for rango in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rango))

    # De abajo hacia arriba: cada borrado cambia los índices de lo que sigue.
    ws.delete_rows(FILA_PROTOTIPO + 1, ws.max_row - FILA_PROTOTIPO)
    ws.delete_rows(FILA_SPEC, 1)
    ws.delete_rows(1, DESPLAZAMIENTO)

    for rango in rangos:
        rango.shift(row_shift=-DESPLAZAMIENTO)
        ws.merge_cells(rango.coord)
    return rangos


def _estilizar_header(ws: Worksheet, ultima_col: int) -> None:
    """Negrita blanca sobre azul en las tres filas de encabezado.

    `merge_cells` sólo copia los bordes de la celda superior izquierda al
    resto del rango combinado: hay que estilizar cada celda, `MergedCell`
    incluidas, para que el relleno y la fuente se vean en toda la combinación.
    """
    for fila in range(1, 4):
        for col in range(1, ultima_col + 1):
            celda = ws.cell(fila, col)
            celda.font = Font(**FUENTE_HEADER)
            celda.fill = RELLENO
            celda.border = BORDE
            celda.alignment = ALINEACION


def _preparar_prototipo(ws: Worksheet) -> None:
    """Vacía la fila 4 y le deja sólo el estilo que usará como prototipo.

    `plantilla_excel._prototipos` lee esta fila entera, así que toda columna
    con estilo debe conservarlo (el valor y el comentario de Contasis, no).
    """
    for celda in ws[4]:
        celda.value = None
        celda.comment = None
        if celda.has_style:
            fuente = copy(celda.font)
            fuente.name, fuente.sz = FUENTE_DATOS
            celda.font = fuente


def _normalizar_columnas(ws: Worksheet) -> None:
    """Separa las definiciones de columna agrupadas en una por columna.

    La plantilla declara anchos por rango (`<col min="1" max="2" .../>`).
    Ensanchar sólo una columna del grupo sin partirlo antes deja dos `<col>`
    superpuestos en el XML y Excel se niega a abrir el archivo sin "reparar".
    El grupo de cola (hasta la última columna de la hoja) se deja intacto.
    """
    dims = ws.column_dimensions
    nuevos = []
    for letra, dim in list(dims.items()):
        if dim.min and dim.max and dim.max > dim.min and dim.max < ULTIMA_COL_TAIL:
            for idx in range(dim.min, dim.max + 1):
                nueva = ColumnDimension(
                    ws, index=get_column_letter(idx), width=dim.width, hidden=dim.hidden
                )
                nueva._style = copy(dim._style)
                nuevos.append(nueva)
            del dims[letra]
    for nueva in nuevos:
        dims[nueva.index] = nueva

    for letra in COLUMNAS_A_MOSTRAR.get(ws.title, ()):
        if letra in dims:
            dims[letra].hidden = False


def _palabra_mas_larga(texto: object) -> int:
    return max((len(palabra) for palabra in str(texto).split()), default=0)


def _ensanchar(ws: Worksheet, ultima_col: int) -> list[tuple[str, object, float, str]]:
    """Sube cada columna visible a un mínimo legible y al ancho de su palabra más larga.

    Nunca reordena ni renombra columnas: Contasis importa por posición. El
    reparto respeta las combinaciones del encabezado (p. ej. compras
    "G10:I10" INFORMACION PROVEEDOR reparte su necesidad entre G, H e I).
    """
    dims = ws.column_dimensions
    ancho_defecto = ws.sheet_format.defaultColWidth or 11.43
    informe: list[tuple[str, object, float, str]] = []

    def dim(idx: int) -> ColumnDimension:
        letra = get_column_letter(idx)
        d = dims.get(letra)
        if d is None:
            d = dims[letra] = ColumnDimension(ws, index=letra, width=ancho_defecto)
        assert not (d.max and d.max >= ULTIMA_COL_TAIL), f"columna {letra} en el grupo de cola"
        return d

    def visible(idx: int) -> bool:
        return not dim(idx).hidden

    def ancho(idx: int) -> float:
        return dim(idx).width or ancho_defecto

    def poner(idx: int, nuevo: float, motivo: str) -> None:
        d = dim(idx)
        viejo = d.width
        d.width = round(nuevo, 2)
        informe.append((get_column_letter(idx), viejo, d.width, motivo))

    for idx in range(1, ultima_col + 1):
        if visible(idx) and ancho(idx) < ANCHO_MINIMO:
            poner(idx, ANCHO_MINIMO, "ancho minimo")

    rangos_por_inicio = {(r.min_row, r.min_col): r for r in ws.merged_cells.ranges}
    celdas_header = []
    for fila in range(1, 4):
        for col in range(1, ultima_col + 1):
            celda = ws.cell(fila, col)
            if isinstance(celda, MergedCell) or celda.value is None:
                continue
            rango = rangos_por_inicio.get((fila, col)) or CellRange(
                min_row=fila, min_col=col, max_row=fila, max_col=col
            )
            celdas_header.append((celda, rango))
    # Los rangos de una sola columna primero: si dos encabezados comparten
    # columnas visibles, quien pide menos no debe robarle ancho a quien pide más.
    celdas_header.sort(key=lambda par: par[1].max_col - par[1].min_col)

    for celda, rango in celdas_header:
        columnas_visibles = [i for i in range(rango.min_col, rango.max_col + 1) if visible(i)]
        if not columnas_visibles:
            continue
        necesario = _palabra_mas_larga(celda.value) + RELLENO_PALABRA
        total = sum(ancho(i) for i in columnas_visibles)
        if total < necesario:
            objetivo = min(columnas_visibles, key=ancho)
            poner(
                objetivo,
                ancho(objetivo) + (necesario - total),
                f"encabezado {celda.coordinate}",
            )
    return informe


def _lineas_necesarias(texto: object, chars_por_linea: int) -> int:
    """Cuántas líneas ocupa `texto` envuelto a `chars_por_linea`, palabra a palabra."""
    lineas, actual = 1, 0
    for palabra in str(texto).split():
        largo = len(palabra)
        if actual == 0:
            actual = largo
        elif actual + 1 + largo <= chars_por_linea:
            actual += 1 + largo
        else:
            lineas += 1
            actual = largo
        while actual > chars_por_linea and chars_por_linea > 0:
            lineas += 1
            actual -= chars_por_linea
    return lineas


def _revisar_alturas(ws: Worksheet, ultima_col: int) -> list[tuple[str, str, int, float, float]]:
    """Encabezados que no caben con `ALTURAS_HEADER`. No falla: sólo avisa."""
    dims = ws.column_dimensions
    ancho_defecto = ws.sheet_format.defaultColWidth or 11.43

    def visible(idx: int) -> bool:
        d = dims.get(get_column_letter(idx))
        return not (d and d.hidden)

    def ancho(idx: int) -> float:
        d = dims.get(get_column_letter(idx))
        return d.width if d and d.width else ancho_defecto

    rangos_por_inicio = {(r.min_row, r.min_col): r for r in ws.merged_cells.ranges}
    problemas = []
    for fila in range(1, 4):
        for col in range(1, ultima_col + 1):
            celda = ws.cell(fila, col)
            if isinstance(celda, MergedCell) or celda.value is None:
                continue
            rango = rangos_por_inicio.get((fila, col)) or CellRange(
                min_row=fila, min_col=col, max_row=fila, max_col=col
            )
            columnas_visibles = [i for i in range(rango.min_col, rango.max_col + 1) if visible(i)]
            if not columnas_visibles:
                continue
            chars = int(sum(ancho(i) for i in columnas_visibles) - 1)
            necesita = _lineas_necesarias(celda.value, chars) * PUNTOS_POR_LINEA + 2
            banda = sum(ALTURAS_HEADER[f] for f in range(rango.min_row, rango.max_row + 1))
            if necesita > banda:
                problemas.append(
                    (celda.coordinate, str(celda.value)[:50], chars, round(necesita, 1), banda)
                )
    return problemas


def _ajustar_vista(ws: Worksheet) -> None:
    ws.row_dimensions.clear()
    for fila, alto in ALTURAS_HEADER.items():
        ws.row_dimensions[fila].height = alto
    ws.sheet_format.defaultRowHeight = ALTURA_DATOS
    ws.sheet_format.customHeight = True

    ws.freeze_panes = "A4"
    ws.sheet_view.zoomScale = 100
    ws.sheet_view.showGridLines = True
    ws.sheet_view.selection[0].activeCell = "A4"
    ws.sheet_view.selection[0].sqref = "A4"
    ws.auto_filter.ref = None
    ws.conditional_formatting = ConditionalFormattingList()


def _transformar_hoja(ws: Worksheet) -> dict:
    ultima_col = _ultima_columna_header(ws)
    rangos = _recortar_filas(ws)
    _estilizar_header(ws, ultima_col)
    _preparar_prototipo(ws)
    _normalizar_columnas(ws)
    anchos = _ensanchar(ws, ultima_col)
    alturas = _revisar_alturas(ws, ultima_col)
    _ajustar_vista(ws)
    return {
        "rangos": [r.coord for r in rangos],
        "ultima_col": ultima_col,
        "anchos": anchos,
        "alturas": alturas,
    }


def preparar(origen: Path, destino: Path) -> dict[str, dict]:
    if not origen.exists():
        raise FileNotFoundError(
            f"no existe {origen}. Es la plantilla oficial de Contasis, fuera del "
            "control de versiones: pídesela a quien la tenga y colócala ahí, o pasa "
            "--origen con su ruta."
        )
    wb = load_workbook(origen)
    resumen = {ws.title: _transformar_hoja(ws) for ws in wb.worksheets}
    wb.active = 0
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)
    return resumen


def _verificar(destino: Path, resumen: dict[str, dict]) -> None:
    """Falla ruidosamente antes de que una plantilla rota llegue a `app/resources`."""
    wb = load_workbook(destino)
    assert set(wb.sheetnames) == set(resumen), wb.sheetnames

    with zipfile.ZipFile(destino) as z:
        nombres_xml = sorted(
            n for n in z.namelist() if re.fullmatch(r"xl/worksheets/sheet\d+\.xml", n)
        )
        xml_por_hoja = dict(
            zip(
                (ws.title for ws in wb.worksheets),
                (z.read(n).decode() for n in nombres_xml),
                strict=True,
            )
        )

    for ws in wb.worksheets:
        info = resumen[ws.title]
        assert ws.max_row == 4, (ws.title, ws.max_row)
        assert str(ws["A1"].value).startswith("FECHA DE EMISION"), (ws.title, ws["A1"].value)
        assert ws["C3"].value == "TIPO", (ws.title, ws["C3"].value)

        combinaciones = sorted(str(r) for r in ws.merged_cells.ranges)
        assert combinaciones == sorted(info["rangos"]), (ws.title, combinaciones)
        assert all(r.max_row <= 3 for r in ws.merged_cells.ranges), ws.title

        formulas = [c.coordinate for fila in ws.iter_rows() for c in fila if c.data_type == "f"]
        assert not formulas, (ws.title, formulas)
        assert all(c.value is None for c in ws[4]), (
            ws.title,
            [c.coordinate for c in ws[4] if c.value],
        )

        assert ws.freeze_panes == "A4", (ws.title, ws.freeze_panes)
        assert ws.sheet_view.zoomScale == 100
        assert ws.auto_filter.ref is None, ws.title
        assert list(ws.conditional_formatting) == [], ws.title
        assert ws.sheet_format.defaultRowHeight == ALTURA_DATOS
        assert {r: d.height for r, d in ws.row_dimensions.items() if d.height} == ALTURAS_HEADER

        anchos_finales = ws.column_dimensions
        estrechas = [
            (letra, d.width)
            for letra, d in anchos_finales.items()
            if not d.hidden
            and d.width
            and d.width < ANCHO_MINIMO
            and (not d.max or d.max < ULTIMA_COL_TAIL)
        ]
        assert not estrechas, (ws.title, estrechas)

        xml = xml_por_hoja[ws.title]
        cols = sorted(
            (int(a), int(b)) for a, b in re.findall(r'<col [^>]*?min="(\d+)" max="(\d+)"', xml)
        )
        for (a1, b1), (a2, b2) in zip(cols, cols[1:], strict=False):
            assert b1 < a2, (ws.title, "columnas <col> superpuestas", (a1, b1), (a2, b2))
        for fila in range(1, 4):
            bloque = re.search(rf'<row r="{fila}"[^>]*>(.*?)</row>', xml, re.S)
            celdas = re.findall(r'<c r="[A-Z]+\d+"([^>]*)', bloque.group(1)) if bloque else []
            sin_estilo = [c for c in celdas if " s=" not in c]
            assert len(celdas) >= info["ultima_col"] and not sin_estilo, (
                ws.title,
                fila,
                len(celdas),
                info["ultima_col"],
            )


def _imprimir_resumen(resumen: dict[str, dict]) -> None:
    for hoja, info in resumen.items():
        print(
            f"\n{hoja}: encabezado hasta columna {get_column_letter(info['ultima_col'])}, "
            f"{len(info['rangos'])} combinaciones"
        )
        if info["anchos"]:
            print("  anchos ajustados:")
            for letra, viejo, nuevo, motivo in info["anchos"]:
                print(f"    {letra:>3}: {viejo!s:>6} -> {nuevo:<6} ({motivo})")
        if info["alturas"]:
            print("  encabezados que podrian no caber con las alturas actuales:")
            for celda, texto, chars, necesita, banda in info["alturas"]:
                print(
                    f'    {celda}: "{texto}" '
                    f"(~{chars} car./linea, necesita {necesita}pt, hay {banda}pt)"
                )


def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument(
        "--origen", type=Path, default=ORIGEN_DEFECTO, help="plantilla oficial de Contasis"
    )
    ap.add_argument("--destino", type=Path, default=DESTINO_DEFECTO, help="archivo a generar")
    args = ap.parse_args()

    resumen = preparar(args.origen, args.destino)
    _verificar(args.destino, resumen)
    _imprimir_resumen(resumen)
    print(f"\nguardado y verificado: {args.destino} ({args.destino.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
