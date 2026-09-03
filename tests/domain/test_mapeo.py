from datetime import date
from decimal import Decimal

from bson.decimal128 import Decimal128
from openpyxl import load_workbook

from app.domain.comprobante import Comprobante, Libro, Origen
from app.repositories.comprobantes import a_documento, desde_documento
from app.services import export_service, plantilla_excel
from app.services.comprobante_service import serializar, texto_para_ia
from app.services.sunat.propuesta import a_comprobante, pertenece_al_periodo

PAYLOAD_SIRE = {
    "numSerieCDP": "F001",
    "numCDP": "123",
    "codTipoCDP": "01",
    "numDocIdentidadProveedor": "20129646099",
    "codTipoDocIdentidadProveedor": "6",
    "desRazonSocialProveedor": "ELECTROCENTRO S.A.",
    "fecEmision": "2026-06-15",
    "codMoneda": "PEN",
    # Forma real del bloque `montos` del RCE, capturada de una respuesta de
    # SUNAT. El fixture anterior usaba `mtoBIGravada`/`mtoIGV`, nombres que el
    # SIRE no envía nunca: los tests pasaban mientras en producción la base y
    # el IGV llegaban siempre en cero.
    "montos": {
        "mtoBIGravadaDG": 100.0,
        "mtoIgvIpmDG": 18.0,
        "mtoBIGravadaDGNG": 0.0,
        "mtoIgvIpmDGNG": 0.0,
        "mtoBIGravadaDNG": 0.0,
        "mtoIgvIpmDNG": 0.0,
        "mtoValorAdqNG": 0.0,
        "mtoIcbp": 0.0,
        "mtoOtrosTrib": 0.0,
        "mtoISC": 0.0,
        "mtoIMB": 0.0,
        "mtoTotalCp": 118.0,
        "mtoBIGravadaDGOriginal": None,
        "mtoIgvIpmDGOriginal": None,
    },
}


# Forma real del RVIE, con los nombres tomados de una respuesta de
# `/rvie/propuesta/web/propuesta/{periodo}/comprobantes`. Ventas separa
# exonerado de inafecto, manda los descuentos aparte y —a diferencia del RCE—
# no trae vencimiento ni tasa de IGV. Los datos del cliente van anonimizados.
PAYLOAD_RVIE = {
    "numSerieCDP": "F001",
    "numCDP": "123",
    "codTipoCDP": "01",
    "codTipoDocIdentidad": "6",
    "numDocIdentidad": "20129646099",
    "nomRazonSocialCliente": "ELECTROCENTRO S.A.",
    # `nomRazonSocial` es la razón social de la propia empresa emisora, no la
    # del cliente. Va en el fixture justo para que se note si alguien vuelve a
    # tomarla por contraparte.
    "nomRazonSocial": "EMPRESA QUE EMITE S.R.L.",
    "fecEmision": "15/06/2026",
    "codMoneda": "PEN",
    "mtoTipoCambio": 1,
    "mtoValFactExpo": 0.0,
    "mtoBIGravada": 100.0,
    "mtoDsctoBI": 0.0,
    "mtoIGV": 18.0,
    "mtoDsctoIGV": 0.0,
    "mtoExonerado": 0.0,
    "mtoInafecto": 0.0,
    "mtoISC": 0.0,
    "mtoBIIvap": 0.0,
    "mtoIvap": 0.0,
    "mtoIcbp": 0.0,
    "mtoOtrosTrib": 0.0,
    "mtoTotalCP": 118.0,
    "codCar": "2012964609901F0010000000123",
    "codEstadoComprobante": "1",
    "indTipoOperacion": "0101",
    "documentoMod": [],
}


class TestMapeoDesdeSire:
    def test_campos_principales(self):
        c = a_comprobante(PAYLOAD_SIRE, Libro.COMPRAS)
        assert c.tipo_cp == "01"
        assert c.serie == "F001"
        assert c.numero == "123"
        assert c.documento_contraparte == "20129646099"
        assert c.razon_social == "ELECTROCENTRO S.A."
        assert c.fecha_emision == date(2026, 6, 15)

    def test_montos_salen_del_bloque_anidado(self):
        c = a_comprobante(PAYLOAD_SIRE, Libro.COMPRAS)
        assert c.base_imponible == Decimal("100.00")
        assert c.igv == Decimal("18.00")
        assert c.total == Decimal("118.00")

    def test_suma_los_tres_destinos_de_la_base_y_del_igv(self):
        # El RCE reparte la base y el IGV entre gravadas (DG), gravadas y no
        # gravadas (DGNG) y no gravadas (DNG). Quedarse con el primero que no
        # sea cero perdería los otros dos.
        payload = {
            **PAYLOAD_SIRE,
            "montos": {
                **PAYLOAD_SIRE["montos"],
                "mtoBIGravadaDG": 100.0,
                "mtoBIGravadaDGNG": 30.0,
                "mtoBIGravadaDNG": 20.0,
                "mtoIgvIpmDG": 18.0,
                "mtoIgvIpmDGNG": 5.4,
                "mtoIgvIpmDNG": 3.6,
            },
        }
        c = a_comprobante(payload, Libro.COMPRAS)
        assert c.base_imponible == Decimal("150.00")
        assert c.igv == Decimal("27.00")

    def test_ignora_los_montos_originales(self):
        # `...Original` guarda el valor previo a una modificación; sumarlo
        # duplicaría la base.
        payload = {
            **PAYLOAD_SIRE,
            "montos": {**PAYLOAD_SIRE["montos"], "mtoBIGravadaDGOriginal": 999.0},
        }
        assert a_comprobante(payload, Libro.COMPRAS).base_imponible == Decimal("100.00")

    def test_otros_tributos_usa_el_nombre_del_rce(self):
        payload = {
            **PAYLOAD_SIRE,
            "montos": {**PAYLOAD_SIRE["montos"], "mtoOtrosTrib": 7.5},
        }
        assert a_comprobante(payload, Libro.COMPRAS).otros_tributos == Decimal("7.50")

    def test_la_fecha_de_vencimiento_llega_en_fec_venc_pag(self):
        # El RCE la manda en `fecVencPag`. Buscándola sólo en `fecVencimiento`
        # o `fecVcto` salía siempre vacía, y el Excel acababa repitiendo la
        # fecha de emisión como si fuera el vencimiento.
        payload = {**PAYLOAD_SIRE, "fecVencPag": "2026-07-31"}
        assert a_comprobante(payload, Libro.COMPRAS).fecha_vencimiento == date(2026, 7, 31)

    def test_sin_vencimiento_el_campo_queda_vacio(self):
        assert a_comprobante(PAYLOAD_SIRE, Libro.COMPRAS).fecha_vencimiento is None

    def test_el_tipo_de_cambio_viene_anidado_en_su_bloque(self):
        # `mtoTipoCambio` no está en la raíz del registro sino dentro de
        # `tipoCambio`; sin aplanar ese bloque las compras en dólares salían
        # sin tipo de cambio.
        payload = {**PAYLOAD_SIRE, "tipoCambio": {"mtoTipoCambio": 3.35}}
        assert a_comprobante(payload, Libro.COMPRAS).tipo_cambio == Decimal("3.35")

    def test_la_tasa_de_igv_se_convierte_a_puntos_porcentuales(self):
        # SUNAT la manda como fracción; el registro la pide como 18 o 10.5.
        payload = {**PAYLOAD_SIRE, "porTasaIGV": 0.105}
        assert a_comprobante(payload, Libro.COMPRAS).porcentaje_igv == Decimal("10.50")

    def test_una_tasa_en_cero_no_es_una_tasa(self):
        # Los comprobantes no gravados llegan con la tasa en cero: dejarla en
        # `None` es lo que permite no escribir un 18 % que no corresponde.
        payload = {**PAYLOAD_SIRE, "porTasaIGV": 0}
        assert a_comprobante(payload, Libro.COMPRAS).porcentaje_igv is None

    def test_desglosa_los_tres_destinos_ademas_de_sumarlos(self):
        payload = {
            **PAYLOAD_SIRE,
            "montos": {
                **PAYLOAD_SIRE["montos"],
                "mtoBIGravadaDG": 100.0,
                "mtoIgvIpmDG": 18.0,
                "mtoBIGravadaDGNG": 50.0,
                "mtoIgvIpmDGNG": 9.0,
                "mtoBIGravadaDNG": 25.0,
                "mtoIgvIpmDNG": 4.5,
            },
        }
        c = a_comprobante(payload, Libro.COMPRAS)
        assert c.base_imponible_dg == Decimal("100.00")
        assert c.base_imponible_dgng == Decimal("50.00")
        assert c.base_imponible_dng == Decimal("25.00")
        assert c.igv_dg == Decimal("18.00")
        assert c.igv_dgng == Decimal("9.00")
        assert c.igv_dng == Decimal("4.50")
        # El total sigue siendo la suma de los tres.
        assert c.base_imponible == Decimal("175.00")
        assert c.igv == Decimal("31.50")

    def test_conserva_el_json_crudo(self):
        c = a_comprobante(PAYLOAD_SIRE, Libro.COMPRAS)
        assert "ELECTROCENTRO" in c.extra["raw_sire"]

    def test_nombre_del_proveedor_gana_al_del_comprador(self):
        # La respuesta del SIRE trae ambos; tomar el equivocado invierte la
        # contraparte del comprobante.
        payload = {
            **PAYLOAD_SIRE,
            "desRazonSocialEmisor": "EMPRESA COMPRADORA SAC",
        }
        assert a_comprobante(payload, Libro.COMPRAS).razon_social == "ELECTROCENTRO S.A."

    def test_filtro_de_periodo_descarta_meses_vecinos(self):
        # SUNAT devuelve comprobantes de periodos adyacentes en la misma página.
        c = a_comprobante(PAYLOAD_SIRE, Libro.COMPRAS)
        assert pertenece_al_periodo(c, "202606")
        assert not pertenece_al_periodo(c, "202605")

    def test_las_boletas_ya_no_se_descartan(self):
        # Antes `serie_aceptada` tiraba todo lo que no empezara con F o E. En
        # ventas eso se llevaba por delante el grueso del libro, así que el
        # filtro desapareció de los dos libros.
        boleta = a_comprobante({**PAYLOAD_SIRE, "numSerieCDP": "B001"}, Libro.COMPRAS)
        assert boleta.serie == "B001"
        assert boleta.es_valido


class TestRoundTripBson:
    def _comprobante(self) -> Comprobante:
        return a_comprobante(PAYLOAD_SIRE, Libro.COMPRAS)

    def test_los_montos_se_guardan_como_decimal128(self):
        documento = a_documento(self._comprobante(), "empresa1", "202606")
        assert isinstance(documento["total"], Decimal128)
        assert isinstance(documento["base_imponible"], Decimal128)

    def test_ida_y_vuelta_conserva_los_valores(self):
        original = self._comprobante()
        recuperado = desde_documento(a_documento(original, "empresa1", "202606"))

        assert recuperado.serie == original.serie
        assert recuperado.numero == original.numero
        assert recuperado.tipo_cp == original.tipo_cp
        assert recuperado.fecha_emision == original.fecha_emision
        assert recuperado.total == original.total
        assert recuperado.igv == original.igv
        assert recuperado.libro is Libro.COMPRAS
        assert recuperado.origen is Origen.SIRE

    def test_el_documento_lleva_serie_numero_derivado(self):
        documento = a_documento(self._comprobante(), "empresa1", "202606")
        assert documento["serie_numero"] == "F001-123"


def _documento_completo() -> dict:
    documento = a_documento(a_comprobante(PAYLOAD_SIRE, Libro.COMPRAS), "empresa1", "202606")
    documento["estado_procesamiento"] = "analizado"
    documento["metadata_procesada"] = {
        "detalle": [
            {
                "producto": "Energía eléctrica",
                "categoria_contable": "Servicios básicos",
                "cantidad": "1",
                "importe": 118.0,
                "razon": "Consumo del local comercial",
            }
        ],
        "cuenta_contable": "6361",
        "resultado": "GASTO",
        "confianza": "95%",
        "observaciones": "Servicio público recurrente",
    }
    return documento


class TestSerializacion:
    def test_los_montos_salen_como_float(self):
        salida = serializar(_documento_completo())
        assert salida["total"] == 118.0
        assert isinstance(salida["total"], float)

    def test_resuelve_la_descripcion_del_tipo(self):
        assert serializar(_documento_completo())["tipo_cp_descripcion"] == "FACTURA"

    def test_expone_el_analisis_con_claves_en_minuscula(self):
        analisis = serializar(_documento_completo())["analisis"]
        assert analisis["resultado"] == "GASTO"
        assert analisis["confianza"] == "95%"

    def test_sin_analisis_devuelve_none(self):
        documento = a_documento(a_comprobante(PAYLOAD_SIRE, Libro.COMPRAS), "e", "202606")
        documento["estado_procesamiento"] = "sire_recibido"
        assert serializar(documento)["analisis"] is None

    def test_texto_para_ia_incluye_lo_normalizado_y_lo_crudo(self):
        texto = texto_para_ia(_documento_completo())
        assert "FACTURA" in texto
        assert "F001-123" in texto
        assert "ELECTROCENTRO" in texto


class TestExportacion:
    def test_excel_de_un_comprobante(self):
        salida = export_service.excel_de_comprobante(serializar(_documento_completo()))
        assert salida.getbuffer().nbytes > 0

    def test_pdf_de_un_comprobante(self):
        salida = export_service.pdf_de_comprobante(serializar(_documento_completo()))
        assert salida.getvalue().startswith(b"%PDF")

    def test_pdf_de_lote(self):
        salida = export_service.pdf_de_lote([serializar(_documento_completo())] * 3)
        assert salida.getvalue().startswith(b"%PDF")


def _documento_ventas() -> dict:
    # Un payload del RVIE, no el del RCE: los dos libros ya no comparten
    # nombres de campo, así que reutilizar el de compras daría todo en cero.
    return a_documento(a_comprobante(PAYLOAD_RVIE, Libro.VENTAS), "empresa1", "202606")


def _hoja(comprobantes: list[dict], libro: Libro):
    salida = plantilla_excel.excel_plantilla(comprobantes, libro)
    wb = load_workbook(salida)
    assert wb.sheetnames == [plantilla_excel.HOJAS[libro]]
    return wb[plantilla_excel.HOJAS[libro]]


class TestPlantillaContasis:
    def test_conserva_los_encabezados_de_la_plantilla(self):
        hoja = _hoja([serializar(_documento_completo())], Libro.COMPRAS)
        assert hoja["A1"].value == "FECHA DE EMISION DEL COMPROBANTE DE PAGO O DOCUMENTO"
        assert hoja["G2"].value == "DOCUMENTO IDENTIDAD"
        assert hoja["C3"].value == "TIPO"

    def test_mapeo_de_compras(self):
        hoja = _hoja([serializar(_documento_completo())], Libro.COMPRAS)
        assert hoja["A4"].value.date() == date(2026, 6, 15)
        assert hoja["C4"].value == "01"
        assert hoja["D4"].value == "F001"
        assert hoja["F4"].value == 123
        assert hoja["G4"].value == 6
        assert hoja["H4"].value == 20129646099
        assert hoja["I4"].value == "ELECTROCENTRO S.A."
        assert hoja["J4"].value == 100.0
        assert hoja["K4"].value == 18.0
        assert hoja["S4"].value == 118.0
        assert hoja["AB4"].value == "S"
        assert hoja["AR4"].value == 18
        # El comprobante está en soles: sin tipo de cambio que declarar.
        assert hoja["W4"].value is None
        assert hoja["AC4"].value is None
        # Sin vencimiento, la condición de pago es al contado.
        assert hoja["AE4"].value == "CON"
        assert hoja["AH4"].value == "4212"

    def test_las_adquisiciones_no_gravadas_llegan_a_la_columna_p(self):
        # El SIRE no separa exonerado de inafecto en compras: manda un único
        # "valor de las adquisiciones no gravadas". La columna P los agrega, y
        # sin contar `no_gravado` salía en cero para toda compra exonerada.
        documento = {**_documento_completo(), "no_gravado": Decimal("59.78")}
        hoja = _hoja([serializar(documento)], Libro.COMPRAS)
        assert hoja["P4"].value == 59.78

    def test_la_columna_p_suma_los_tres_conceptos(self):
        documento = {
            **_documento_completo(),
            "exonerado": Decimal("10.00"),
            "inafecto": Decimal("5.00"),
            "no_gravado": Decimal("20.00"),
        }
        hoja = _hoja([serializar(documento)], Libro.COMPRAS)
        assert hoja["P4"].value == 35.0

    def test_los_tres_destinos_van_a_columnas_distintas(self):
        # J/K son sólo las gravadas; DGNG y DNG tienen sus propias columnas.
        # Mandando la suma a J/K se declaraba como gravado lo destinado a
        # operaciones no gravadas: el total cuadraba, el destino no.
        documento = {
            **_documento_completo(),
            "base_imponible_dg": Decimal("100.00"),
            "igv_dg": Decimal("18.00"),
            "base_imponible_dgng": Decimal("50.00"),
            "igv_dgng": Decimal("9.00"),
            "base_imponible_dng": Decimal("25.00"),
            "igv_dng": Decimal("4.50"),
        }
        hoja = _hoja([serializar(documento)], Libro.COMPRAS)
        assert hoja["J4"].value == 100.0
        assert hoja["K4"].value == 18.0
        assert hoja["L4"].value == 50.0
        assert hoja["M4"].value == 9.0
        assert hoja["N4"].value == 25.0
        assert hoja["O4"].value == 4.5

    def test_el_tipo_de_cambio_llega_a_la_columna_w(self):
        # En soles no hay nada que convertir: hace falta una moneda extranjera
        # para que el tipo de cambio y la conversión entren en juego.
        documento = {
            **_documento_completo(),
            "moneda": "USD",
            "tipo_cambio": Decimal("3.387"),
        }
        hoja = _hoja([serializar(documento)], Libro.COMPRAS)
        assert hoja["W4"].value == 3.387
        assert hoja["AB4"].value == "D"
        # Importes convertidos a soles con el tipo de cambio de la fila.
        assert hoja["J4"].value == 338.7
        assert hoja["K4"].value == 60.97
        assert hoja["S4"].value == 399.67
        # El total original en dólares va aparte, sin convertir.
        assert hoja["AC4"].value == 118.0

    def test_la_tasa_declarada_gana_a_la_general(self):
        documento = {**_documento_completo(), "porcentaje_igv": Decimal("10.50")}
        hoja = _hoja([serializar(documento)], Libro.COMPRAS)
        assert hoja["AR4"].value == 10.5

    def test_sin_igv_la_tasa_sigue_siendo_la_general(self):
        # Los cuatro registros reales usados para comparar esta exportación
        # llevan 18 en todas las filas, también en las que no tienen IGV.
        documento = {
            **_documento_completo(),
            "igv": Decimal("0.00"),
            "igv_dg": Decimal("0.00"),
            "porcentaje_igv": None,
        }
        hoja = _hoja([serializar(documento)], Libro.COMPRAS)
        assert hoja["AR4"].value == plantilla_excel.TASA_IGV

    def test_con_igv_y_sin_tasa_declarada_cae_en_la_general(self):
        documento = {**_documento_completo(), "porcentaje_igv": None}
        hoja = _hoja([serializar(documento)], Libro.COMPRAS)
        assert hoja["AR4"].value == plantilla_excel.TASA_IGV

    def test_mapeo_de_ventas(self):
        hoja = _hoja([serializar(_documento_ventas())], Libro.VENTAS)
        assert hoja["C4"].value == "01"
        assert hoja["D4"].value == "F001"
        assert hoja["E4"].value == 123
        assert hoja["H4"].value == "ELECTROCENTRO S.A."
        assert hoja["J4"].value == 100.0
        assert hoja["N4"].value == 18.0
        assert hoja["P4"].value == 118.0
        assert hoja["V4"].value == "S"
        assert hoja["AL4"].value == 18
        assert hoja["Q4"].value is None
        assert hoja["W4"].value is None
        assert hoja["Y4"].value == "CON"
        assert hoja["AD4"].value == "1212"

    def test_condicion_de_pago_es_credito_con_vencimiento_posterior(self):
        documento = {**_documento_completo(), "fecha_vencimiento": date(2026, 7, 31)}
        hoja = _hoja([serializar(documento)], Libro.COMPRAS)
        assert hoja["AE4"].value == "CRE"

    def test_condicion_de_pago_es_contado_si_el_vencimiento_no_es_posterior(self):
        documento = {**_documento_completo(), "fecha_vencimiento": date(2026, 6, 15)}
        hoja = _hoja([serializar(documento)], Libro.COMPRAS)
        assert hoja["AE4"].value == "CON"

    def test_el_analisis_llena_cuenta_contable_y_glosa(self):
        hoja = _hoja([serializar(_documento_completo())], Libro.COMPRAS)
        assert hoja["AF4"].value == "6361"
        # Siempre en mayúsculas, como la escriben los contadores.
        assert hoja["AS4"].value == "ENERGÍA ELÉCTRICA"

    def test_el_analisis_tambien_viaja_en_la_hoja_de_ventas(self):
        # La hoja de ventas coloca las mismas columnas en otras letras: la
        # moneda es la V, así que AB es la cuenta contable.
        documento = _documento_ventas()
        documento["metadata_procesada"] = {
            "cuenta_contable": "7011",
            "detalle": [{"producto": "Venta de mercadería"}],
        }
        hoja = _hoja([serializar(documento)], Libro.VENTAS)
        assert hoja["AB4"].value == "7011"
        assert hoja["AM4"].value == "VENTA DE MERCADERÍA"

    def test_el_centro_de_costos_no_se_escribe(self):
        # La plantilla pide el código del catálogo de Contasis (9 caracteres) y
        # la IA devuelve un nombre: recortarlo inventaría códigos que además
        # colisionan entre sí ("Administración" y "Administración y Finanzas"
        # darían el mismo).
        documento = _documento_completo()
        documento["metadata_procesada"]["centro_costos"] = "Operaciones - Flota Vehicular"
        hoja = _hoja([serializar(documento)], Libro.COMPRAS)
        assert hoja["AI4"].value is None
        assert hoja["AJ4"].value is None

    def test_el_rag_completa_cuentas_codigos_y_glosa_de_compras(self):
        documento = _documento_completo()
        documento["metadata_procesada"]["rag"] = {
            "codigo_comprobante": "07",
            "codigo_identidad": "1",
            "cuenta_base": "6365095",
            "cuenta_total": "4212",
            "glosa": "SERVICIO DE INTERNET",
        }
        hoja = _hoja([serializar(documento)], Libro.COMPRAS)
        assert hoja["C4"].value == "07"
        assert hoja["G4"].value == 1
        assert hoja["AF4"].value == "6365095"
        assert hoja["AH4"].value == "4212"
        assert hoja["AS4"].value == "SERVICIO DE INTERNET"

    def test_la_cuenta_total_del_rag_gana_a_la_general_del_libro(self):
        # El RAG puede resolver una cuenta propia de la empresa, distinta de
        # la 4212/1212 general: si la trae, es la que se escribe.
        documento = _documento_completo()
        documento["metadata_procesada"]["rag"] = {"cuenta_total": "42121"}
        hoja = _hoja([serializar(documento)], Libro.COMPRAS)
        assert hoja["AH4"].value == "42121"

    def test_las_cuentas_que_la_ia_no_deduce_siguen_vacias(self):
        # La cuenta de otros tributos no la da ni el análisis ni el RAG; la
        # del total sí tiene valor siempre, aunque nadie la haya deducido: cae
        # en la cuenta general del libro.
        hoja = _hoja([serializar(_documento_completo())], Libro.COMPRAS)
        assert hoja["AG4"].value is None
        assert hoja["AH4"].value == "4212"

    def test_la_cuenta_total_en_ventas_es_la_general_sin_rag(self):
        hoja = _hoja([serializar(_documento_ventas())], Libro.VENTAS)
        assert hoja["AD4"].value == "1212"

    def test_la_glosa_se_recorta_al_ancho_de_la_columna(self):
        documento = _documento_completo()
        documento["metadata_procesada"]["detalle"] = [
            {"producto": "Servicio de mantenimiento preventivo y correctivo de la flota"},
            {"producto": "Repuestos varios"},
        ]
        documento["metadata_procesada"]["descripcion"] = (
            "Adquisición de servicios de mantenimiento preventivo y correctivo "
            "para la flota vehicular de la empresa durante el periodo"
        )
        hoja = _hoja([serializar(documento)], Libro.COMPRAS)
        glosa = hoja["AS4"].value
        assert len(glosa) <= plantilla_excel.MAX_GLOSA
        # Cortada por palabra, no a mitad de una.
        assert glosa == "ADQUISICIÓN DE SERVICIOS DE MANTENIMIENTO PREVENTIVO Y"

    def test_con_un_solo_item_la_glosa_es_el_producto(self):
        documento = _documento_completo()
        documento["metadata_procesada"]["descripcion"] = "Resumen largo del comprobante"
        hoja = _hoja([serializar(documento)], Libro.COMPRAS)
        assert hoja["AS4"].value == "ENERGÍA ELÉCTRICA"

    def test_con_varios_items_la_glosa_es_el_resumen(self):
        # El nombre del primer ítem describiría sólo una parte de la compra.
        documento = _documento_completo()
        documento["metadata_procesada"]["detalle"] = [
            {"producto": "Aceite de motor"},
            {"producto": "Filtro de aire"},
        ]
        documento["metadata_procesada"]["descripcion"] = "Insumos de mantenimiento"
        hoja = _hoja([serializar(documento)], Libro.COMPRAS)
        assert hoja["AS4"].value == "INSUMOS DE MANTENIMIENTO"

    def test_sin_analisis_las_columnas_de_la_ia_quedan_vacias(self):
        documento = _documento_completo()
        documento["metadata_procesada"] = None
        hoja = _hoja([serializar(documento)], Libro.COMPRAS)
        assert hoja["AF4"].value is None
        assert hoja["AS4"].value is None

    def test_la_referencia_de_la_nota_de_credito_llega_a_ventas(self):
        # Sólo el RVIE manda `documentoMod`; el RCE no lo hace nunca, así que
        # esta referencia sólo existe en la hoja de ventas.
        payload = {
            **PAYLOAD_RVIE,
            "codTipoCDP": "07",
            "documentoMod": [
                {"codTipoCDP": "03", "numSerieCDP": "B001", "numCDP": "5777"}
            ],
        }
        documento = a_documento(a_comprobante(payload, Libro.VENTAS), "empresa1", "202606")
        hoja = _hoja([serializar(documento)], Libro.VENTAS)
        assert hoja["S4"].value == "03"
        assert hoja["T4"].value == "B001"
        assert hoja["U4"].value == 5777
        assert hoja["R4"].value is None

    def test_la_fecha_de_la_nota_de_credito_se_escribe_si_viene(self):
        payload = {
            **PAYLOAD_RVIE,
            "codTipoCDP": "07",
            "documentoMod": [
                {
                    "codTipoCDP": "03",
                    "numSerieCDP": "B001",
                    "numCDP": "5777",
                    "fecEmision": "02/05/2025",
                }
            ],
        }
        documento = a_documento(a_comprobante(payload, Libro.VENTAS), "empresa1", "202606")
        hoja = _hoja([serializar(documento)], Libro.VENTAS)
        assert hoja["R4"].value.date() == date(2025, 5, 2)
        assert hoja["R4"].number_format == plantilla_excel.FORMATO_FECHA

    def test_compras_no_tiene_referencia_a_documento_modificado(self):
        # El RCE nunca manda `documentoMod`: esas columnas se quedan vacías.
        hoja = _hoja([serializar(_documento_completo())], Libro.COMPRAS)
        assert [hoja[c + "4"].value for c in ("X", "Y", "Z", "AA")] == [None, None, None, None]

    def test_pie_de_totales_suma_el_rango_de_datos(self):
        # Un único pie: todo el registro ya está en soles.
        hoja = _hoja([serializar(_documento_completo())] * 3, Libro.COMPRAS)
        assert hoja["A7"].value == "TOTAL"
        assert hoja["S7"].value == "=SUM(S4:S6)"
        assert hoja["A8"].value is None

    def test_sin_comprobantes_salen_solo_los_encabezados(self):
        hoja = _hoja([], Libro.VENTAS)
        assert hoja.max_row == 3
        assert hoja["A1"].value == "FECHA DE EMISION DEL COMPROBANTE DE PAGO O DOCUMENTO"

    def test_las_filas_de_ejemplo_de_la_plantilla_no_sobreviven(self):
        # La plantilla no trae filas de ejemplo por debajo del prototipo, pero
        # si alguna vez las trajera, sólo debería sobrevivir el pie de totales.
        hoja = _hoja([serializar(_documento_completo())], Libro.COMPRAS)
        formulas = [
            str(c.value)
            for fila in hoja.iter_rows(min_row=4)
            for c in fila
            if isinstance(c.value, str) and c.value.startswith("=")
        ]
        assert all(f.startswith("=SUM(") for f in formulas)

    def test_las_columnas_que_llena_la_exportacion_no_quedan_ocultas(self):
        # La plantilla oficial trae INAFECTA e ISC ocultas en ventas; esta
        # exportación sí las llena, así que no pueden quedar invisibles.
        hoja = _hoja([serializar(_documento_ventas())], Libro.VENTAS)
        assert hoja.column_dimensions["L"].hidden is False
        assert hoja.column_dimensions["M"].hidden is False
        assert hoja.column_dimensions["O"].hidden is False

    def test_las_filas_no_llevan_una_altura_propia(self):
        # La altura uniforme la resuelve la plantilla (`defaultRowHeight`); si
        # `openpyxl` dejara alturas sueltas de una plantilla vieja, las filas
        # de datos y el pie se verían más chicas o más grandes que el resto.
        hoja = _hoja([serializar(_documento_completo())] * 2, Libro.COMPRAS)
        assert hoja.sheet_format.defaultRowHeight == 18.75
        assert all(hoja.row_dimensions[fila].height is None for fila in (4, 5, 6))
